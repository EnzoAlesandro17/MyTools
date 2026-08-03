"""MyTools - servidor local (Flask + SQLite).

Sirve el shell unico (app.html) + los modulos estaticos, y expone la API
que usan modules/arqueo-caja/arqueo.js y modules/control-fibra/fibra.js.
Pensado para correr solo en localhost via run.sh / run.bat.
"""
import io
import json
import re
import sqlite3
import time
import urllib.error
import urllib.parse
import urllib.request
import uuid
from pathlib import Path

from flask import Flask, g, jsonify, request, send_file, send_from_directory
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / 'data'
DB_PATH = DATA_DIR / 'mytools.db'
SCHEMA_PATH = BASE_DIR / 'schema.sql'

app = Flask(__name__, static_folder=None)

EPS = 0.005


# ---------------------------------------------------------------- DB ----

def get_db():
    db = getattr(g, '_db', None)
    if db is None:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        db = g._db = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
        db.execute('PRAGMA foreign_keys = ON')
    return db


@app.teardown_appcontext
def close_db(_exc):
    db = getattr(g, '_db', None)
    if db is not None:
        db.close()


def migrate_db(conn):
    """Agrega columnas y estructuras nuevas a bases ya existentes (CREATE
    TABLE IF NOT EXISTS no las suma solo). Cada paso se aplica una sola vez,
    comprobando antes si hace falta - seguro de correr en cada arranque."""
    cols = {r['name'] for r in conn.execute('PRAGMA table_info(empleados)')}
    for col in ('apellido', 'dni', 'telefono', 'email', 'codigo_interno'):
        if col not in cols:
            conn.execute(f'ALTER TABLE empleados ADD COLUMN {col} TEXT')

    horario_cols = {r['name'] for r in conn.execute('PRAGMA table_info(horarios)')}
    if 'semana' not in horario_cols:
        conn.execute('ALTER TABLE horarios ADD COLUMN semana TEXT')
    if 'sin_turno' not in horario_cols:
        conn.execute('ALTER TABLE horarios ADD COLUMN sin_turno INTEGER NOT NULL DEFAULT 0')
    if 'tipo' not in horario_cols:
        conn.execute('ALTER TABLE horarios ADD COLUMN tipo TEXT')
    if 'nota' not in horario_cols:
        conn.execute('ALTER TABLE horarios ADD COLUMN nota TEXT')
    conn.execute('CREATE INDEX IF NOT EXISTS idx_horarios_semana ON horarios(semana)')

    # Backfill de 'tipo' para filas cargadas con versiones anteriores (no
    # tenian este campo). Las filas viejas de "plantilla general" (semana
    # NULL) quedan como estaban, sin usarse - el modelo nuevo ya no las lee.
    conn.execute("UPDATE horarios SET tipo = 'franco' WHERE tipo IS NULL AND sin_turno = 1")
    conn.execute('''
        UPDATE horarios SET tipo = 'horario'
        WHERE tipo IS NULL AND hora_inicio IS NOT NULL AND hora_inicio != ''
          AND hora_fin IS NOT NULL AND hora_fin != ''
    ''')

    # El modelo nuevo permite un solo registro por empleado/dia/semana. Las
    # versiones anteriores permitian turno cortado (2 filas para el mismo
    # dia). Antes de crear el indice unico, fusionamos esos pares en una sola
    # fila tipo 'otro' con el detalle en texto, para no perder el dato ni
    # romper la migracion.
    grupos = conn.execute('''
        SELECT empleado_id, dia_semana, semana FROM horarios
        WHERE semana IS NOT NULL
        GROUP BY empleado_id, dia_semana, semana
        HAVING COUNT(*) > 1
    ''').fetchall()
    for g in grupos:
        filas = conn.execute(
            'SELECT * FROM horarios WHERE empleado_id=? AND dia_semana=? AND semana=? ORDER BY hora_inicio',
            (g['empleado_id'], g['dia_semana'], g['semana']),
        ).fetchall()
        partes = [f"{f['hora_inicio']}\u2013{f['hora_fin']}" for f in filas if f['hora_inicio'] and f['hora_fin']]
        conn.execute(
            "UPDATE horarios SET tipo='otro', hora_inicio=NULL, hora_fin=NULL, nota=? WHERE id=?",
            (' y '.join(partes), filas[0]['id']),
        )
        for f in filas[1:]:
            conn.execute('DELETE FROM horarios WHERE id=?', (f['id'],))

    conn.execute(
        'CREATE UNIQUE INDEX IF NOT EXISTS uq_horarios_emp_dia_semana ON horarios(empleado_id, dia_semana, semana)'
    )

    conn.commit()


def init_db():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA_PATH.read_text(encoding='utf-8'))
    migrate_db(conn)
    conn.close()


# ------------------------------------------------------------ helpers ----

def new_id():
    return uuid.uuid4().hex[:16]


def now_ms():
    return int(time.time() * 1000)


# Formato argentino: "." separa miles (grupos de 3 digitos exactos), "," es
# el separador decimal. Un token sin punto de miles cae al caso simple
# (numero llano, coma decimal opcional) para no romper "700000" o "+100000+50000".
NUMBER_TOKEN_RE = re.compile(r'[+-]?\d{1,3}(?:\.\d{3})+(?:,\d+)?|[+-]?\d+(?:,\d+)?')


def eval_expr(expr):
    if expr is None:
        return 0.0
    cleaned = re.sub(r'\s+', '', str(expr))
    if not cleaned:
        return 0.0
    matches = NUMBER_TOKEN_RE.findall(cleaned)
    if not matches:
        return 0.0
    return sum(float(tok.replace('.', '').replace(',', '.')) for tok in matches)


def to_field(expr):
    e = '0' if expr in (None, '') else str(expr).strip()
    return {'expr': e, 'val': eval_expr(e)}


def bad_request(msg):
    return jsonify({'error': msg}), 400


# --------------------------------------------------------- static app ----

@app.get('/')
def index():
    return send_from_directory(BASE_DIR, 'app.html')


@app.get('/shared/<path:filename>')
def shared_files(filename):
    return send_from_directory(BASE_DIR / 'shared', filename)


@app.get('/modules/<path:filename>')
def module_files(filename):
    return send_from_directory(BASE_DIR / 'modules', filename)


# ===========================================================================
# Sucursales (se gestionan desde Configuracion > General)
# ===========================================================================

def row_to_sucursal(row):
    return {
        'id': row['id'], 'nombre': row['nombre'],
        'codigoInterno': row['codigo_interno'] or '', 'direccion': row['direccion'] or '',
    }


@app.get('/api/sucursales')
def sucursales_list():
    db = get_db()
    rows = db.execute('SELECT * FROM sucursales ORDER BY nombre COLLATE NOCASE').fetchall()
    return jsonify([row_to_sucursal(r) for r in rows])


@app.post('/api/sucursales')
def sucursales_create():
    data = request.get_json(force=True, silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    if not nombre:
        return bad_request('nombre requerido')
    codigo_interno = (data.get('codigoInterno') or '').strip()
    direccion = (data.get('direccion') or '').strip()
    db = get_db()
    sid = new_id()
    db.execute(
        'INSERT INTO sucursales (id, nombre, codigo_interno, direccion) VALUES (?, ?, ?, ?)',
        (sid, nombre, codigo_interno, direccion),
    )
    db.commit()
    row = db.execute('SELECT * FROM sucursales WHERE id = ?', (sid,)).fetchone()
    return jsonify(row_to_sucursal(row)), 201


@app.delete('/api/sucursales/<sid>')
def sucursales_delete(sid):
    db = get_db()
    db.execute('DELETE FROM sucursales WHERE id = ?', (sid,))
    db.commit()
    return '', 204


# ===========================================================================
# Empleados (compartido - se gestionan desde Configuracion > General)
# ===========================================================================

def row_to_empleado(row):
    return {
        'id': row['id'], 'nombre': row['nombre'], 'apellido': row['apellido'] or '',
        'dni': row['dni'] or '', 'telefono': row['telefono'] or '', 'email': row['email'] or '',
        'codigoInterno': row['codigo_interno'] or '',
    }


@app.get('/api/empleados')
def empleados_list():
    db = get_db()
    rows = db.execute('SELECT * FROM empleados ORDER BY nombre COLLATE NOCASE').fetchall()
    return jsonify([row_to_empleado(r) for r in rows])


@app.post('/api/empleados')
def empleados_create():
    data = request.get_json(force=True, silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    apellido = (data.get('apellido') or '').strip()
    dni = (data.get('dni') or '').strip()
    telefono = (data.get('telefono') or '').strip()
    email = (data.get('email') or '').strip()
    codigo_interno = (data.get('codigoInterno') or '').strip()
    if not nombre:
        return bad_request('nombre requerido')
    db = get_db()
    existing = db.execute(
        'SELECT * FROM empleados WHERE nombre = ? COLLATE NOCASE AND apellido = ? COLLATE NOCASE',
        (nombre, apellido),
    ).fetchone()
    if existing:
        return jsonify(row_to_empleado(existing))
    eid = new_id()
    db.execute(
        '''INSERT INTO empleados (id, nombre, apellido, dni, telefono, email, codigo_interno)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (eid, nombre, apellido, dni, telefono, email, codigo_interno),
    )
    db.commit()
    row = db.execute('SELECT * FROM empleados WHERE id = ?', (eid,)).fetchone()
    return jsonify(row_to_empleado(row)), 201


@app.delete('/api/empleados/<eid>')
def empleados_delete(eid):
    db = get_db()
    db.execute('DELETE FROM horario_empleados WHERE empleado_id = ?', (eid,))
    db.execute('DELETE FROM horarios WHERE empleado_id = ?', (eid,))
    db.execute('DELETE FROM empleados WHERE id = ?', (eid,))
    db.commit()
    return '', 204


# ===========================================================================
# Arqueo de caja
# ===========================================================================

def row_to_arqueo(db, row):
    emp_rows = db.execute(
        'SELECT nombre FROM arqueo_empleados WHERE arqueo_id = ?', (row['id'],)
    ).fetchall()
    return {
        'id': row['id'],
        'fecha': row['fecha'],
        'empleados': [r['nombre'] for r in emp_rows],
        'cf': {'expr': row['cf_expr'], 'val': row['cf_val']},
        'cc': {'expr': row['cc_expr'], 'val': row['cc_val']},
        'sc': {'expr': row['sc_expr'], 'val': row['sc_val']},
        'resultado': row['resultado'],
        'eliminado': bool(row['eliminado']),
    }


@app.get('/api/arqueo/registros')
def arqueo_registros_list():
    db = get_db()
    rows = db.execute('SELECT * FROM arqueos WHERE eliminado = 0 ORDER BY fecha DESC').fetchall()
    return jsonify([row_to_arqueo(db, r) for r in rows])


def save_arqueo_empleados(db, arqueo_id, empleados):
    db.execute('DELETE FROM arqueo_empleados WHERE arqueo_id = ?', (arqueo_id,))
    for nombre in empleados or []:
        nombre = (nombre or '').strip()
        if nombre:
            db.execute(
                'INSERT INTO arqueo_empleados (arqueo_id, nombre) VALUES (?, ?)',
                (arqueo_id, nombre),
            )


@app.post('/api/arqueo/registros')
def arqueo_registros_create():
    data = request.get_json(force=True, silent=True) or {}
    fecha = (data.get('fecha') or '').strip()
    empleados = data.get('empleados') or []
    if not fecha or not empleados:
        return bad_request('fecha y empleados son requeridos')
    cf = to_field(data.get('cf'))
    cc = to_field(data.get('cc'))
    sc = to_field(data.get('sc'))
    resultado = cf['val'] + cc['val'] - sc['val']
    db = get_db()
    aid = new_id()
    db.execute(
        '''INSERT INTO arqueos (id, fecha, cf_expr, cf_val, cc_expr, cc_val, sc_expr, sc_val, resultado, eliminado)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)''',
        (aid, fecha, cf['expr'], cf['val'], cc['expr'], cc['val'], sc['expr'], sc['val'], resultado),
    )
    save_arqueo_empleados(db, aid, empleados)
    db.commit()
    row = db.execute('SELECT * FROM arqueos WHERE id = ?', (aid,)).fetchone()
    return jsonify(row_to_arqueo(db, row)), 201


@app.put('/api/arqueo/registros/<aid>')
def arqueo_registros_update(aid):
    db = get_db()
    existing = db.execute('SELECT * FROM arqueos WHERE id = ?', (aid,)).fetchone()
    if not existing:
        return bad_request('no existe')
    data = request.get_json(force=True, silent=True) or {}
    fecha = (data.get('fecha') or '').strip()
    empleados = data.get('empleados') or []
    if not fecha or not empleados:
        return bad_request('fecha y empleados son requeridos')
    cf = to_field(data.get('cf'))
    cc = to_field(data.get('cc'))
    sc = to_field(data.get('sc'))
    resultado = cf['val'] + cc['val'] - sc['val']
    db.execute(
        '''UPDATE arqueos SET fecha=?, cf_expr=?, cf_val=?, cc_expr=?, cc_val=?, sc_expr=?, sc_val=?, resultado=?
           WHERE id=?''',
        (fecha, cf['expr'], cf['val'], cc['expr'], cc['val'], sc['expr'], sc['val'], resultado, aid),
    )
    save_arqueo_empleados(db, aid, empleados)
    db.commit()
    row = db.execute('SELECT * FROM arqueos WHERE id = ?', (aid,)).fetchone()
    return jsonify(row_to_arqueo(db, row))


@app.delete('/api/arqueo/registros/<aid>')
def arqueo_registros_delete(aid):
    db = get_db()
    db.execute('UPDATE arqueos SET eliminado = 1 WHERE id = ?', (aid,))
    db.commit()
    return '', 204


ARQUEO_EXCEL_HEADERS = ['Fecha y hora', 'Empleados', 'Caja fuerte', 'Caja chica', 'Saldo sistema', 'Resultado']


def fmt_fecha_hora_larga(iso):
    m = re.match(r'^(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})', iso or '')
    if not m:
        return ''
    y, mo, d, h, mi = m.groups()
    return f'{d}/{mo}/{y} {h}:{mi}'


@app.get('/api/arqueo/export.xlsx')
def arqueo_export_xlsx():
    db = get_db()
    rows = db.execute(
        'SELECT * FROM arqueos WHERE eliminado = 0 ORDER BY fecha'
    ).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Arqueos'
    ws.append(ARQUEO_EXCEL_HEADERS)
    for row in rows:
        arqueo = row_to_arqueo(db, row)
        ws.append([
            fmt_fecha_hora_larga(arqueo['fecha']),
            ', '.join(arqueo['empleados']),
            arqueo['cf']['val'],
            arqueo['cc']['val'],
            arqueo['sc']['val'],
            arqueo['resultado'],
        ])
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'arqueo_db-{time.strftime("%Y-%m-%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def parse_arqueo_fecha(val):
    import datetime
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime('%Y-%m-%dT%H:%M:%S.000000')
    s = str(val).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})[ T](\d{1,2}):(\d{2})', s)
    if not m:
        return None
    d, mo, y, h, mi = m.groups()
    return f'{y}-{mo.zfill(2)}-{d.zfill(2)}T{h.zfill(2)}:{mi}:00.000000'


@app.post('/api/arqueo/import-excel')
def arqueo_import_excel():
    file = request.files.get('file')
    if not file:
        return bad_request('falta el archivo')
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        return bad_request('no se pudo leer el Excel')
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    except StopIteration:
        return bad_request('sin filas')
    if 'Fecha y hora' not in headers:
        return bad_request('Tiene que tener las mismas columnas que genera "Exportar Excel".')

    db = get_db()
    added = 0
    for raw_row in rows_iter:
        row = dict(zip(headers, raw_row))
        fecha_iso = parse_arqueo_fecha(row.get('Fecha y hora')) if row.get('Fecha y hora') else None
        if not fecha_iso:
            continue
        empleados = [n.strip() for n in str(row.get('Empleados') or '').split(',') if n.strip()]
        cf = to_field(row.get('Caja fuerte'))
        cc = to_field(row.get('Caja chica'))
        sc = to_field(row.get('Saldo sistema'))
        resultado = cf['val'] + cc['val'] - sc['val']
        aid = new_id()
        db.execute(
            '''INSERT INTO arqueos (id, fecha, cf_expr, cf_val, cc_expr, cc_val, sc_expr, sc_val, resultado, eliminado)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 0)''',
            (aid, fecha_iso, cf['expr'], cf['val'], cc['expr'], cc['val'], sc['expr'], sc['val'], resultado),
        )
        save_arqueo_empleados(db, aid, empleados)
        for nombre in empleados:
            if not db.execute('SELECT id FROM empleados WHERE nombre = ? COLLATE NOCASE', (nombre,)).fetchone():
                db.execute('INSERT INTO empleados (id, nombre) VALUES (?, ?)', (new_id(), nombre))
        added += 1

    db.commit()
    return jsonify({'added': added})


# ===========================================================================
# Control de fibra
# ===========================================================================

FIELD_IDS = [
    'vendedor', 'fechaIngreso', 'plan', 'cantidadTV', 'fechaPactada', 'franjaPactada', 'ot', 'sds',
    'fechaInstalacion', 'estado', 'observaciones', 'clienteNombre', 'dni', 'fechaNacimiento', 'email',
    'telefono', 'telefonoAlt', 'localidad', 'calle', 'altura', 'entreCalles', 'tipoDomicilio', 'torrePisoDepto',
]
FIELD_TO_COL = {
    'vendedor': 'vendedor', 'fechaIngreso': 'fecha_ingreso', 'plan': 'plan', 'cantidadTV': 'cantidad_tv',
    'fechaPactada': 'fecha_pactada', 'franjaPactada': 'franja_pactada', 'ot': 'ot', 'sds': 'sds',
    'fechaInstalacion': 'fecha_instalacion', 'estado': 'estado', 'observaciones': 'observaciones',
    'clienteNombre': 'cliente_nombre', 'dni': 'dni', 'fechaNacimiento': 'fecha_nacimiento', 'email': 'email',
    'telefono': 'telefono', 'telefonoAlt': 'telefono_alt', 'localidad': 'localidad', 'calle': 'calle',
    'altura': 'altura', 'entreCalles': 'entre_calles', 'tipoDomicilio': 'tipo_domicilio',
    'torrePisoDepto': 'torre_piso_depto',
}
EXCEL_HEADERS = [
    'Vendedor', 'Fecha ingreso', 'Plan (MB)', 'Cant. TV', 'Fecha pactada', 'Franja', 'OT', 'SDS',
    'Fecha instalación', 'Estado', 'Con formulario', 'Observaciones', 'Titular', 'DNI/CUIT', 'Fecha nacimiento',
    'Mail', 'Teléfono', 'Alternativo', 'Localidad', 'Calle', 'Altura', 'Entre calles', 'Tipo domicilio',
    'Torre/Piso/Depto',
]
EXCEL_KEYS = [
    'vendedor', 'fechaIngreso', 'plan', 'cantidadTV', 'fechaPactada', 'franjaPactada', 'ot', 'sds',
    'fechaInstalacion', 'estado', 'conForm', 'observaciones', 'clienteNombre', 'dni', 'fechaNacimiento', 'email',
    'telefono', 'telefonoAlt', 'localidad', 'calle', 'altura', 'entreCalles', 'tipoDomicilio', 'torrePisoDepto',
]
DATE_KEYS = {'fechaIngreso', 'fechaPactada', 'fechaInstalacion', 'fechaNacimiento'}


@app.get('/api/fibra/planes')
def fibra_planes_list():
    db = get_db()
    rows = db.execute('SELECT * FROM planes ORDER BY mb').fetchall()
    return jsonify([{'id': r['id'], 'mb': r['mb']} for r in rows])


@app.post('/api/fibra/planes')
def fibra_planes_create():
    data = request.get_json(force=True, silent=True) or {}
    try:
        mb = int(data.get('mb'))
    except (TypeError, ValueError):
        return bad_request('mb invalido')
    if mb <= 0:
        return bad_request('mb invalido')
    db = get_db()
    existing = db.execute('SELECT * FROM planes WHERE mb = ?', (mb,)).fetchone()
    if existing:
        return jsonify({'id': existing['id'], 'mb': existing['mb']})
    pid = new_id()
    db.execute('INSERT INTO planes (id, mb) VALUES (?, ?)', (pid, mb))
    db.commit()
    return jsonify({'id': pid, 'mb': mb}), 201


@app.delete('/api/fibra/planes/<pid>')
def fibra_planes_delete(pid):
    db = get_db()
    db.execute('DELETE FROM planes WHERE id = ?', (pid,))
    db.commit()
    return '', 204


def row_to_venta(row):
    out = {'id': row['id'], 'conForm': bool(row['con_form']), 'eliminado': bool(row['eliminado']),
           'orden': row['orden'], 'createdAt': row['created_at'], 'updatedAt': row['updated_at']}
    for field_id, col in FIELD_TO_COL.items():
        out[field_id] = row[col]
    return out


def venta_values_from_payload(data):
    values = {}
    for field_id in FIELD_IDS:
        values[field_id] = (data.get(field_id) or '').strip() if isinstance(data.get(field_id), str) else (data.get(field_id) or '')
    return values


@app.get('/api/fibra/ventas')
def fibra_ventas_list():
    db = get_db()
    rows = db.execute(
        'SELECT * FROM ventas WHERE eliminado = 0 ORDER BY fecha_ingreso DESC, orden ASC, created_at ASC'
    ).fetchall()
    return jsonify([row_to_venta(r) for r in rows])


@app.post('/api/fibra/ventas')
def fibra_ventas_create():
    data = request.get_json(force=True, silent=True) or {}
    if not (data.get('vendedor') and data.get('fechaIngreso') and data.get('clienteNombre')):
        return bad_request('vendedor, fechaIngreso y clienteNombre son requeridos')
    values = venta_values_from_payload(data)
    con_form = bool(data.get('conForm'))
    db = get_db()
    vid = new_id()
    ts = now_ms()
    min_orden = db.execute(
        'SELECT MIN(orden) AS m FROM ventas WHERE eliminado = 0 AND fecha_ingreso = ?',
        (values['fechaIngreso'],),
    ).fetchone()['m']
    orden = (min_orden - 10) if min_orden is not None else 10
    cols = ', '.join(FIELD_TO_COL[f] for f in FIELD_IDS)
    placeholders = ', '.join('?' for _ in FIELD_IDS)
    db.execute(
        f'''INSERT INTO ventas (id, {cols}, con_form, orden, created_at, updated_at, eliminado)
            VALUES (?, {placeholders}, ?, ?, ?, ?, 0)''',
        (vid, *[values[f] for f in FIELD_IDS], 1 if con_form else 0, orden, ts, ts),
    )
    db.commit()
    row = db.execute('SELECT * FROM ventas WHERE id = ?', (vid,)).fetchone()
    return jsonify(row_to_venta(row)), 201


@app.put('/api/fibra/ventas/<vid>')
def fibra_ventas_update(vid):
    db = get_db()
    existing = db.execute('SELECT * FROM ventas WHERE id = ?', (vid,)).fetchone()
    if not existing:
        return bad_request('no existe')
    data = request.get_json(force=True, silent=True) or {}
    if not (data.get('vendedor') and data.get('fechaIngreso') and data.get('clienteNombre')):
        return bad_request('vendedor, fechaIngreso y clienteNombre son requeridos')
    values = venta_values_from_payload(data)
    con_form = bool(data.get('conForm'))
    assignments = ', '.join(f'{FIELD_TO_COL[f]} = ?' for f in FIELD_IDS)
    db.execute(
        f'UPDATE ventas SET {assignments}, con_form = ?, updated_at = ? WHERE id = ?',
        (*[values[f] for f in FIELD_IDS], 1 if con_form else 0, now_ms(), vid),
    )
    db.commit()
    row = db.execute('SELECT * FROM ventas WHERE id = ?', (vid,)).fetchone()
    return jsonify(row_to_venta(row))


@app.delete('/api/fibra/ventas/<vid>')
def fibra_ventas_delete(vid):
    db = get_db()
    db.execute('UPDATE ventas SET eliminado = 1, updated_at = ? WHERE id = ?', (now_ms(), vid))
    db.commit()
    return '', 204


@app.put('/api/fibra/ventas/reorder')
def fibra_ventas_reorder():
    data = request.get_json(force=True, silent=True) or {}
    drag_id, target_id = data.get('dragId'), data.get('targetId')
    if not drag_id or not target_id or drag_id == target_id:
        return '', 204
    db = get_db()
    drag = db.execute('SELECT * FROM ventas WHERE id = ? AND eliminado = 0', (drag_id,)).fetchone()
    target = db.execute('SELECT * FROM ventas WHERE id = ? AND eliminado = 0', (target_id,)).fetchone()
    if not drag or not target or drag['fecha_ingreso'] != target['fecha_ingreso']:
        return '', 204
    day = drag['fecha_ingreso']
    day_rows = db.execute(
        'SELECT id FROM ventas WHERE eliminado = 0 AND fecha_ingreso = ? ORDER BY orden ASC, created_at ASC',
        (day,),
    ).fetchall()
    ids = [r['id'] for r in day_rows]
    from_idx, to_idx = ids.index(drag_id), ids.index(target_id)
    ids.insert(to_idx, ids.pop(from_idx))
    ts = now_ms()
    for i, vid in enumerate(ids):
        db.execute('UPDATE ventas SET orden = ?, updated_at = ? WHERE id = ?', ((i + 1) * 10, ts, vid))
    db.commit()
    return '', 204


@app.get('/api/fibra/export.xlsx')
def fibra_export_xlsx():
    db = get_db()
    rows = db.execute(
        'SELECT * FROM ventas WHERE eliminado = 0 ORDER BY fecha_ingreso, orden'
    ).fetchall()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas'
    ws.append(EXCEL_HEADERS)
    for row in rows:
        v = row_to_venta(row)
        line = []
        for key in EXCEL_KEYS:
            if key == 'conForm':
                line.append('Sí' if v['conForm'] else 'No')
            elif key in ('plan', 'cantidadTV'):
                line.append(int(v[key]) if v.get(key) else '')
            else:
                line.append(v.get(key) or '')
        ws.append(line)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f'ventas-fibra-{time.strftime("%Y-%m-%d")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


def capitalize_words(s):
    if not s:
        return s
    return ' '.join(w[:1].upper() + w[1:] if w else w for w in s.lower().split(' '))


def excel_val_to_date_str(val):
    if val in (None, ''):
        return ''
    import datetime
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = ('20' if int(y) < 50 else '19') + y
        return f'{y}-{mo.zfill(2)}-{d.zfill(2)}'
    return ''


def excel_val_to_bool(val):
    s = str(val if val is not None else '').strip().lower()
    return s in ('sí', 'si', 'true', '1')


@app.post('/api/fibra/import-excel')
def fibra_import_excel():
    file = request.files.get('file')
    if not file:
        return bad_request('falta el archivo')
    try:
        wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        return bad_request('no se pudo leer el Excel')
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else '' for h in next(rows_iter)]
    except StopIteration:
        return bad_request('sin filas')
    header_to_key = {h: k for h, k in zip(EXCEL_HEADERS, EXCEL_KEYS)}

    db = get_db()
    added = 0
    ts_base = now_ms()
    for raw_row in rows_iter:
        row = dict(zip(headers, raw_row))
        rec = {}
        for header, val in row.items():
            key = header_to_key.get(header)
            if not key:
                continue
            if key in DATE_KEYS:
                val = excel_val_to_date_str(val)
            elif key == 'conForm':
                val = excel_val_to_bool(val)
            elif key in ('plan', 'cantidadTV'):
                try:
                    val = str(int(float(val))) if val not in (None, '') else ''
                except (TypeError, ValueError):
                    val = ''
            else:
                val = str(val).strip() if val is not None else ''
            rec[key] = val
        if not rec.get('vendedor') and not rec.get('clienteNombre'):
            continue
        rec['vendedor'] = capitalize_words(rec.get('vendedor', ''))
        rec['clienteNombre'] = capitalize_words(rec.get('clienteNombre', ''))
        rec['calle'] = capitalize_words(rec.get('calle', ''))
        rec['localidad'] = capitalize_words(rec.get('localidad', ''))
        rec['entreCalles'] = capitalize_words(rec.get('entreCalles', ''))
        rec['torrePisoDepto'] = capitalize_words(rec.get('torrePisoDepto', ''))
        rec['ot'] = (rec.get('ot') or '').upper()
        rec['sds'] = (rec.get('sds') or '').upper()
        rec['email'] = (rec.get('email') or '').lower()

        values = venta_values_from_payload(rec)
        vid = new_id()
        ts = ts_base + added
        cols = ', '.join(FIELD_TO_COL[f] for f in FIELD_IDS)
        placeholders = ', '.join('?' for _ in FIELD_IDS)
        db.execute(
            f'''INSERT INTO ventas (id, {cols}, con_form, orden, created_at, updated_at, eliminado)
                VALUES (?, {placeholders}, ?, ?, ?, ?, 0)''',
            (vid, *[values[f] for f in FIELD_IDS], 1 if rec.get('conForm') else 0, ts, ts, ts),
        )
        if rec['vendedor'] and not db.execute(
            'SELECT id FROM empleados WHERE nombre = ? COLLATE NOCASE', (rec['vendedor'],)
        ).fetchone():
            db.execute('INSERT INTO empleados (id, nombre) VALUES (?, ?)', (new_id(), rec['vendedor']))
        if rec.get('plan'):
            try:
                mb = int(rec['plan'])
            except ValueError:
                mb = 0
            if mb and not db.execute('SELECT id FROM planes WHERE mb = ?', (mb,)).fetchone():
                db.execute('INSERT INTO planes (id, mb) VALUES (?, ?)', (new_id(), mb))
        added += 1

    db.commit()
    return jsonify({'added': added})


# ===========================================================================
# Enlaces utiles
# ===========================================================================

def normalize_url(url):
    url = (url or '').strip()
    if url and not re.match(r'^[a-zA-Z][a-zA-Z0-9+.-]*://', url):
        url = 'https://' + url
    return url


def row_to_enlace(row):
    return {
        'id': row['id'], 'titulo': row['titulo'], 'url': row['url'],
        'categoria': row['categoria'] or '', 'descripcion': row['descripcion'] or '',
        'createdAt': row['created_at'],
    }


@app.get('/api/enlaces')
def enlaces_list():
    db = get_db()
    rows = db.execute(
        'SELECT * FROM enlaces ORDER BY categoria COLLATE NOCASE, titulo COLLATE NOCASE'
    ).fetchall()
    return jsonify([row_to_enlace(r) for r in rows])


@app.post('/api/enlaces')
def enlaces_create():
    data = request.get_json(force=True, silent=True) or {}
    titulo = (data.get('titulo') or '').strip()
    url = normalize_url(data.get('url'))
    if not titulo or not url:
        return bad_request('titulo y url son requeridos')
    db = get_db()
    eid = new_id()
    db.execute(
        'INSERT INTO enlaces (id, titulo, url, categoria, descripcion, created_at) VALUES (?, ?, ?, ?, ?, ?)',
        (eid, titulo, url, (data.get('categoria') or '').strip(), (data.get('descripcion') or '').strip(), now_ms()),
    )
    db.commit()
    row = db.execute('SELECT * FROM enlaces WHERE id = ?', (eid,)).fetchone()
    return jsonify(row_to_enlace(row)), 201


@app.put('/api/enlaces/<eid>')
def enlaces_update(eid):
    db = get_db()
    existing = db.execute('SELECT id FROM enlaces WHERE id = ?', (eid,)).fetchone()
    if not existing:
        return bad_request('no existe')
    data = request.get_json(force=True, silent=True) or {}
    titulo = (data.get('titulo') or '').strip()
    url = normalize_url(data.get('url'))
    if not titulo or not url:
        return bad_request('titulo y url son requeridos')
    db.execute(
        'UPDATE enlaces SET titulo=?, url=?, categoria=?, descripcion=? WHERE id=?',
        (titulo, url, (data.get('categoria') or '').strip(), (data.get('descripcion') or '').strip(), eid),
    )
    db.commit()
    row = db.execute('SELECT * FROM enlaces WHERE id = ?', (eid,)).fetchone()
    return jsonify(row_to_enlace(row))


@app.delete('/api/enlaces/<eid>')
def enlaces_delete(eid):
    db = get_db()
    db.execute('DELETE FROM enlaces WHERE id = ?', (eid,))
    db.commit()
    return '', 204


# ===========================================================================
# Tareas
# ===========================================================================

TAREA_COMMENT_PREVIEW = 5


def row_to_comentario(row):
    return {'id': row['id'], 'texto': row['texto'], 'createdAt': row['created_at']}


def row_to_tarea(db, row):
    comentarios = db.execute(
        'SELECT * FROM tarea_comentarios WHERE tarea_id = ? ORDER BY created_at DESC LIMIT ?',
        (row['id'], TAREA_COMMENT_PREVIEW),
    ).fetchall()
    total = db.execute(
        'SELECT count(*) c FROM tarea_comentarios WHERE tarea_id = ?', (row['id'],)
    ).fetchone()['c']
    return {
        'id': row['id'], 'titulo': row['titulo'], 'fechaLimite': row['fecha_limite'] or '',
        'prioritaria': bool(row['prioritaria']), 'estado': row['estado'],
        'createdAt': row['created_at'], 'closedAt': row['closed_at'],
        'comentarios': [row_to_comentario(c) for c in comentarios],
        'comentariosCount': total,
    }


@app.get('/api/tareas')
def tareas_list():
    db = get_db()
    rows = db.execute('SELECT * FROM tareas ORDER BY created_at ASC').fetchall()
    return jsonify([row_to_tarea(db, r) for r in rows])


@app.post('/api/tareas')
def tareas_create():
    data = request.get_json(force=True, silent=True) or {}
    titulo = (data.get('titulo') or '').strip()
    if not titulo:
        return bad_request('titulo requerido')
    fecha_limite = (data.get('fechaLimite') or '').strip()
    prioritaria = bool(data.get('prioritaria'))
    db = get_db()
    tid = new_id()
    db.execute(
        '''INSERT INTO tareas (id, titulo, fecha_limite, prioritaria, estado, created_at)
           VALUES (?, ?, ?, ?, 'abierta', ?)''',
        (tid, titulo, fecha_limite, 1 if prioritaria else 0, now_ms()),
    )
    db.commit()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    return jsonify(row_to_tarea(db, row)), 201


@app.put('/api/tareas/<tid>/prioridad')
def tareas_toggle_prioridad(tid):
    db = get_db()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    if not row:
        return bad_request('no existe')
    db.execute('UPDATE tareas SET prioritaria = ? WHERE id = ?', (0 if row['prioritaria'] else 1, tid))
    db.commit()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    return jsonify(row_to_tarea(db, row))


@app.post('/api/tareas/<tid>/cerrar')
def tareas_cerrar(tid):
    db = get_db()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    if not row:
        return bad_request('no existe')
    if row['estado'] == 'cerrada':
        return bad_request('la tarea ya esta cerrada')
    db.execute("UPDATE tareas SET estado = 'cerrada', closed_at = ? WHERE id = ?", (now_ms(), tid))
    db.commit()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    return jsonify(row_to_tarea(db, row))


@app.post('/api/tareas/<tid>/reabrir')
def tareas_reabrir(tid):
    db = get_db()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    if not row:
        return bad_request('no existe')
    if row['estado'] != 'cerrada':
        return bad_request('la tarea no esta cerrada')
    db.execute("UPDATE tareas SET estado = 'abierta', closed_at = NULL WHERE id = ?", (tid,))
    db.commit()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    return jsonify(row_to_tarea(db, row))


@app.get('/api/tareas/<tid>/comentarios')
def tarea_comentarios_list(tid):
    db = get_db()
    rows = db.execute(
        'SELECT * FROM tarea_comentarios WHERE tarea_id = ? ORDER BY created_at DESC', (tid,)
    ).fetchall()
    return jsonify([row_to_comentario(r) for r in rows])


@app.post('/api/tareas/<tid>/comentarios')
def tarea_comentarios_create(tid):
    db = get_db()
    tarea = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    if not tarea:
        return bad_request('no existe')
    if tarea['estado'] == 'cerrada':
        return bad_request('la tarea esta cerrada')
    data = request.get_json(force=True, silent=True) or {}
    texto = (data.get('texto') or '').strip()
    if not texto:
        return bad_request('texto requerido')
    cid = new_id()
    db.execute(
        'INSERT INTO tarea_comentarios (id, tarea_id, texto, created_at) VALUES (?, ?, ?, ?)',
        (cid, tid, texto, now_ms()),
    )
    db.commit()
    row = db.execute('SELECT * FROM tareas WHERE id = ?', (tid,)).fetchone()
    return jsonify(row_to_tarea(db, row)), 201


# ===========================================================================
# Horarios
# ===========================================================================

DIAS_SEMANA_COUNT = 7
TIPOS_HORARIO = ('horario', 'licencia', 'franco', 'otro')


def row_to_horario(row):
    return {
        'id': row['id'], 'empleadoId': row['empleado_id'], 'diaSemana': row['dia_semana'],
        'semana': row['semana'], 'tipo': row['tipo'] or '',
        'horaInicio': row['hora_inicio'] or '', 'horaFin': row['hora_fin'] or '',
        'nota': row['nota'] or '',
    }


@app.get('/api/horarios/semana/<semana>')
def horarios_por_semana(semana):
    db = get_db()
    rows = db.execute('SELECT * FROM horarios WHERE semana = ? ORDER BY dia_semana', (semana,)).fetchall()
    return jsonify([row_to_horario(r) for r in rows])


@app.put('/api/horarios/dia')
def horarios_set_dia():
    """Fija el horario de un empleado para un dia y una semana especifica (no
    hay plantilla general: cada semana se carga entera). tipo:
      - 'horario' -> requiere horaInicio/horaFin.
      - 'otro'    -> requiere nota (texto libre).
      - 'licencia' / 'franco' -> sin datos adicionales."""
    data = request.get_json(force=True, silent=True) or {}
    empleado_id = (data.get('empleadoId') or '').strip()
    if not empleado_id:
        return bad_request('empleado requerido')
    db = get_db()
    empleado = db.execute('SELECT id FROM empleados WHERE id = ?', (empleado_id,)).fetchone()
    if not empleado:
        return bad_request('el empleado no existe')
    try:
        dia_semana = int(data.get('diaSemana'))
    except (TypeError, ValueError):
        return bad_request('dia invalido')
    if dia_semana < 0 or dia_semana >= DIAS_SEMANA_COUNT:
        return bad_request('dia invalido')

    semana = (data.get('semana') or '').strip()
    if not re.match(r'^\d{4}-W\d{2}$', semana):
        return bad_request('semana invalida')

    tipo = (data.get('tipo') or '').strip()
    if tipo not in TIPOS_HORARIO:
        return bad_request('tipo invalido')

    hora_inicio = hora_fin = nota = None
    if tipo == 'horario':
        hora_inicio = (data.get('horaInicio') or '').strip()
        hora_fin = (data.get('horaFin') or '').strip()
        if not re.match(r'^\d{2}:\d{2}$', hora_inicio) or not re.match(r'^\d{2}:\d{2}$', hora_fin):
            return bad_request('hora invalida')
        if hora_fin <= hora_inicio:
            return bad_request('la hora de fin debe ser posterior a la de inicio')
    elif tipo == 'otro':
        nota = (data.get('nota') or '').strip()
        if not nota:
            return bad_request('escribi un detalle para "Otros"')

    db.execute(
        '''INSERT INTO horarios (id, empleado_id, dia_semana, semana, tipo, hora_inicio, hora_fin, nota)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(empleado_id, dia_semana, semana) DO UPDATE SET
             tipo=excluded.tipo, hora_inicio=excluded.hora_inicio, hora_fin=excluded.hora_fin, nota=excluded.nota''',
        (new_id(), empleado_id, dia_semana, semana, tipo, hora_inicio, hora_fin, nota),
    )
    db.commit()
    row = db.execute(
        'SELECT * FROM horarios WHERE empleado_id = ? AND dia_semana = ? AND semana = ?',
        (empleado_id, dia_semana, semana),
    ).fetchone()
    return jsonify(row_to_horario(row))


@app.delete('/api/horarios/dia')
def horarios_delete_dia():
    empleado_id = (request.args.get('empleadoId') or '').strip()
    semana = (request.args.get('semana') or '').strip()
    try:
        dia_semana = int(request.args.get('diaSemana'))
    except (TypeError, ValueError):
        return bad_request('dia invalido')
    if not empleado_id or not semana:
        return bad_request('empleado y semana requeridos')
    db = get_db()
    db.execute(
        'DELETE FROM horarios WHERE empleado_id = ? AND dia_semana = ? AND semana = ?',
        (empleado_id, dia_semana, semana),
    )
    db.commit()
    return '', 204


# ---------- Orden y visibilidad de empleados en la grilla de Horarios ----------

def sync_horario_empleados(db):
    """Da de alta en horario_empleados a cualquier empleado que todavia no
    tenga fila (nuevos en Configuracion), agregandolos visibles al final."""
    empleado_ids = [r['id'] for r in db.execute('SELECT id FROM empleados').fetchall()]
    existentes = {r['empleado_id'] for r in db.execute('SELECT empleado_id FROM horario_empleados').fetchall()}
    siguiente = db.execute('SELECT COALESCE(MAX(orden), -1) + 1 AS n FROM horario_empleados').fetchone()['n']
    nuevos = False
    for eid in empleado_ids:
        if eid not in existentes:
            db.execute(
                'INSERT INTO horario_empleados (empleado_id, orden, visible) VALUES (?, ?, 1)',
                (eid, siguiente),
            )
            siguiente += 1
            nuevos = True
    if nuevos:
        db.commit()


@app.get('/api/horarios/empleados')
def horario_empleados_list():
    db = get_db()
    sync_horario_empleados(db)
    rows = db.execute('SELECT * FROM horario_empleados ORDER BY orden').fetchall()
    return jsonify([
        {'empleadoId': r['empleado_id'], 'orden': r['orden'], 'visible': bool(r['visible'])}
        for r in rows
    ])


@app.put('/api/horarios/empleados')
def horario_empleados_set():
    """Reemplaza orden y visibilidad para todos los empleados incluidos en
    'items'. Se manda la lista completa desde el frontend cada vez."""
    data = request.get_json(force=True, silent=True) or {}
    items = data.get('items')
    if not isinstance(items, list):
        return bad_request('items invalido')
    db = get_db()
    for it in items:
        eid = (it.get('empleadoId') or '').strip()
        if not eid:
            continue
        try:
            orden = int(it.get('orden'))
        except (TypeError, ValueError):
            continue
        visible = 1 if it.get('visible') else 0
        db.execute(
            '''INSERT INTO horario_empleados (empleado_id, orden, visible) VALUES (?, ?, ?)
               ON CONFLICT(empleado_id) DO UPDATE SET orden=excluded.orden, visible=excluded.visible''',
            (eid, orden, visible),
        )
    db.commit()
    rows = db.execute('SELECT * FROM horario_empleados ORDER BY orden').fetchall()
    return jsonify([
        {'empleadoId': r['empleado_id'], 'orden': r['orden'], 'visible': bool(r['visible'])}
        for r in rows
    ])


# ===========================================================================
# Configuracion global + clima
# ===========================================================================

CONFIG_DEFAULTS = {
    'nombre_negocio': 'MyTools',
    'clima_ciudad': 'Rosario',
    'clima_lat': '-32.9468',
    'clima_lon': '-60.6393',
    'clima_unidad_temp': 'C',
    'clima_unidad_presion': 'hPa',
}

WEATHER_CODES = {
    0: ('Despejado', '☀️'), 1: ('Mayormente despejado', '🌤️'), 2: ('Parcialmente nublado', '⛅'),
    3: ('Nublado', '☁️'), 45: ('Niebla', '🌫️'), 48: ('Niebla', '🌫️'),
    51: ('Llovizna débil', '🌦️'), 53: ('Llovizna', '🌦️'), 55: ('Llovizna intensa', '🌦️'),
    56: ('Llovizna helada', '🌦️'), 57: ('Llovizna helada intensa', '🌦️'),
    61: ('Lluvia débil', '🌧️'), 63: ('Lluvia', '🌧️'), 65: ('Lluvia intensa', '🌧️'),
    66: ('Lluvia helada', '🌧️'), 67: ('Lluvia helada intensa', '🌧️'),
    71: ('Nieve débil', '🌨️'), 73: ('Nieve', '🌨️'), 75: ('Nieve intensa', '🌨️'), 77: ('Granizo fino', '🌨️'),
    80: ('Chubascos débiles', '🌦️'), 81: ('Chubascos', '🌧️'), 82: ('Chubascos intensos', '⛈️'),
    85: ('Chubascos de nieve', '🌨️'), 86: ('Chubascos de nieve intensos', '🌨️'),
    95: ('Tormenta', '⛈️'), 96: ('Tormenta con granizo', '⛈️'), 99: ('Tormenta con granizo', '⛈️'),
}


def get_config_dict(db):
    rows = db.execute('SELECT clave, valor FROM config_global').fetchall()
    values = dict(CONFIG_DEFAULTS)
    values.update({r['clave']: r['valor'] for r in rows if r['valor'] not in (None, '')})
    return values


@app.get('/api/config')
def config_get():
    return jsonify(get_config_dict(get_db()))


@app.put('/api/config')
def config_put():
    data = request.get_json(force=True, silent=True) or {}
    db = get_db()
    for key in CONFIG_DEFAULTS:
        if key in data:
            val = str(data[key]).strip()
            db.execute(
                '''INSERT INTO config_global (clave, valor) VALUES (?, ?)
                   ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor''',
                (key, val),
            )
    db.commit()
    return jsonify(get_config_dict(db))


@app.get('/api/config/info')
def config_info():
    db = get_db()
    counts = {}
    for table in ('sucursales', 'empleados', 'arqueos', 'planes', 'ventas', 'enlaces', 'tareas', 'horarios'):
        counts[table] = db.execute(f'SELECT count(*) c FROM {table}').fetchone()['c']
    size = DB_PATH.stat().st_size if DB_PATH.exists() else 0
    return jsonify({'dbPath': str(DB_PATH), 'dbSizeKb': round(size / 1024, 1), 'counts': counts})


@app.get('/api/weather')
def weather():
    db = get_db()
    cfg = get_config_dict(db)
    lat = request.args.get('lat') or cfg['clima_lat']
    lon = request.args.get('lon') or cfg['clima_lon']
    unidad_temp = cfg.get('clima_unidad_temp') or 'C'
    unidad_presion = cfg.get('clima_unidad_presion') or 'hPa'
    params = {'latitude': lat, 'longitude': lon, 'current': 'temperature_2m,weathercode,surface_pressure'}
    if unidad_temp == 'F':
        params['temperature_unit'] = 'fahrenheit'
    url = f'https://api.open-meteo.com/v1/forecast?{urllib.parse.urlencode(params)}'
    try:
        with urllib.request.urlopen(url, timeout=4) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except (urllib.error.URLError, TimeoutError, ValueError):
        return jsonify({'error': 'clima no disponible'}), 503
    current = data.get('current') or {}
    code = current.get('weathercode')
    desc, icon = WEATHER_CODES.get(code, ('', '🌡️'))
    presion = current.get('surface_pressure')
    if presion is not None and unidad_presion == 'mmHg':
        presion = round(presion * 0.750062, 1)
    return jsonify({
        'temp': current.get('temperature_2m'),
        'unidadTemp': unidad_temp,
        'code': code,
        'desc': desc,
        'icon': icon,
        'presion': presion,
        'unidadPresion': unidad_presion,
        'ciudad': cfg['clima_ciudad'],
    })


@app.get('/api/weather/forecast')
def weather_forecast():
    db = get_db()
    cfg = get_config_dict(db)
    lat = request.args.get('lat') or cfg['clima_lat']
    lon = request.args.get('lon') or cfg['clima_lon']
    unidad_temp = cfg.get('clima_unidad_temp') or 'C'
    params = {
        'latitude': lat, 'longitude': lon, 'timezone': 'auto', 'forecast_days': 7,
        'daily': 'weathercode,temperature_2m_max,temperature_2m_min',
    }
    if unidad_temp == 'F':
        params['temperature_unit'] = 'fahrenheit'
    url = f'https://api.open-meteo.com/v1/forecast?{urllib.parse.urlencode(params)}'
    try:
        with urllib.request.urlopen(url, timeout=4) as resp:
            data = json.loads(resp.read().decode('utf-8'))
    except (urllib.error.URLError, TimeoutError, ValueError):
        return jsonify({'error': 'clima no disponible'}), 503
    daily = data.get('daily') or {}
    fechas = daily.get('time') or []
    codes = daily.get('weathercode') or []
    maxs = daily.get('temperature_2m_max') or []
    mins = daily.get('temperature_2m_min') or []
    dias = []
    for i, fecha in enumerate(fechas):
        code = codes[i] if i < len(codes) else None
        desc, icon = WEATHER_CODES.get(code, ('', '🌡️'))
        dias.append({
            'fecha': fecha,
            'code': code,
            'desc': desc,
            'icon': icon,
            'tempMax': maxs[i] if i < len(maxs) else None,
            'tempMin': mins[i] if i < len(mins) else None,
        })
    return jsonify({'unidadTemp': unidad_temp, 'ciudad': cfg['clima_ciudad'], 'dias': dias})


if __name__ == '__main__':
    init_db()
    app.run(host='127.0.0.1', port=5000, debug=False)
